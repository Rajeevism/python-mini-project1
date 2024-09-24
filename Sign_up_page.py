#Sign Up Page
from tkinter import *
from tkinter import messagebox
import openpyxl

window = Tk()
window.title('Sign Up Page')
window.geometry('1080x600+300+200')
window.configure(bg="#A3D8FF")
window.resizable(True,True)

# Improved error handling and data validation
def sign_up():
    User_name = user_name.get()
    email_id = user_id.get()
    passwd = password.get()
    confirm_passwd = confirmed_password.get()

    if not all([User_name,email_id, passwd, confirm_passwd]):  # Check if all fields are filled
        messagebox.showerror("Invalid", "Please fill in all fields.")
        return

    if not valid_email(email_id):  # Validate email format
        messagebox.showerror("Invalid", "Please enter a valid email address.")
        return

    if passwd != confirm_passwd:
        messagebox.showerror("Invalid", "Both passwords should match.")
        return

    try:
        # Efficient data handling with openpyxl
        workbook = openpyxl.load_workbook('user_data.xlsx')
        sheet = workbook.active  # Assuming the first sheet
        # Check for existing email before writing to Excel
        for row in sheet.iter_rows(min_row=2):
          if row[1].value == email_id:  # Check email in second column (index 1)
            messagebox.showerror("Error", "You have already signed up with this email address.")
            return
        
        max_row = sheet.max_row + 1  # Get the next available row
        sheet.cell(row=max_row, column=1).value = User_name
        sheet.cell(row=max_row, column=2).value = email_id
        sheet.cell(row=max_row, column=3).value = passwd
        workbook.save('user_data.xlsx')
        messagebox.showinfo('Sign Up', "Successfully signed up!")

    except FileNotFoundError:
        # Create the Excel file if it doesn't exist
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1).value = 'User Name'
        sheet.cell(row=1, column=2).value = 'Email Id'
        sheet.cell(row=1, column=3).value = 'Password'
        sheet.cell(row=2, column=1).value = User_name
        sheet.cell(row=2, column=2).value = email_id
        sheet.cell(row=2, column=3).value = passwd
        workbook.save('user_data.xlsx')
#------------------------------------------------------------------
def valid_email(email):
    # Basic email validation (can be enhanced)
    return '@' in email and '.' in email
#---------------------------------------------------------------------
    
img = PhotoImage(file='sign_up_icon.png')
Label(window,image=img,bg='#A3D8FF').place(x=50,y=70)

frame = Frame(window,width=600,height=500,bg='#A3D8FF')
frame.place(x=480,y=70)

heading = Label(frame,text='Sign Up',fg='blue',bg='#A3D8FF',font=('Times New Roman',28,'bold'))
heading.place(x=150,y=5)

####------------------------------------------
def on_enter(e):
      user_name.delete(0,'end')
def on_leave(e):
      Username=user_name.get()
      if Username==' ':
            user_name.insert(0,'Enter your name')
user_name = Entry(frame,width=35,fg='blue',border=1,bg='#A3D8FF',font=('Times New Roman',16))
user_name.place(x=40,y=80)
user_name.insert(0,'Enter your name')
user_name.bind('<FocusIn>',on_enter)
user_name.bind('<FocusOut>',on_leave)
####------------------------------------------
def on_enter(e):
      user_id.delete(0,'end')
def on_leave(e):
      name=user_id.get()
      if name==' ':
            user_id.insert(0,'Enter your email id')
user_id = Entry(frame,width=35,fg='blue',border=1,bg='#A3D8FF',font=('Times New Roman',16))
user_id.place(x=40,y=120)
user_id.insert(0,'Enter your email id')
user_id.bind('<FocusIn>',on_enter)
user_id.bind('<FocusOut>',on_leave)
####-------------------------------------------
def on_enter(e):
      password.delete(0,'end')
def on_leave(e):
      name1=password.get()
      if name1==' ':
            password.insert(0,'Enter your password')
password = Entry(frame,width=35,fg='blue',border=1,bg='#A3D8FF',font=('Times New Roman',16))
password.place(x=40,y=160)
password.insert(0,'Enter your password')
password.bind('<FocusIn>',on_enter)
password.bind('<FocusOut>',on_leave)
#--------------------------------------------------------------#
def on_enter(e):
      confirmed_password.delete(0,'end')
def on_leave(e):
      name2 = confirmed_password.get()
      if name2==' ':
            confirmed_password.insert(0,'Confirm your password')
confirmed_password = Entry(frame,width=35,fg='blue',border=1,bg='#A3D8FF',font=('Times New Roman',16))
confirmed_password.place(x=40,y=200)
confirmed_password.insert(0,'Confirm your password')
confirmed_password.bind('<FocusIn>',on_enter)
confirmed_password.bind('<FocusOut>',on_leave)
#--------------------------------------------------------------------
def open_sign_in():
    window.destroy()# Destroy the current login window
    import Login_page#importing login page file

#---------------------------------------------------------------------
Button(frame,width=35,pady=7,text='Sign Up', bg='blue',fg='white',border=0,command=sign_up).place(x=35,y=250)
label=Label(frame,text="Already have an account ?",fg='blue',bg='#A3D8FF',font=('Times New Roman',16))
label.place(x=40,y=300)

sign_in = Button(frame,width=10,text='Sign in',border=0,bg='#A3D8FF',cursor='hand2',fg='blue',command=open_sign_in)
sign_in.place(x=260,y=305)
window.mainloop()
